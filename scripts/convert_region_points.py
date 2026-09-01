#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Convert the survey-point workbook into the runtime region_points.json file.

The source workbook may omit the province column; in that case every row is
assigned the default province.  Keeping province in the JSON now allows the
same validation contract to scale to a multi-province catalog later.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = PROJECT_ROOT / "backend" / "data" / "region_points.json"
DEFAULT_PROVINCE = "贵州省"

HEADER_ALIASES = {
    "省": "province",
    "市": "city",
    "市/州": "city",
    "县": "county",
    "县/区": "county",
    "乡镇": "township",
    "乡/镇": "township",
    "村居": "community",
    "社区/村": "community",
}
REQUIRED_HEADERS = ("city", "county", "township", "community")
REGION_FIELDS = ("province", "city", "county", "township", "community")

# Administrative zones that use a parent county in the source workbook but are
# managed independently in the survey system.  Keyed by (city, township).
COUNTY_OVERRIDES: dict[tuple[str, str], str] = {
    ("遵义市", "三渡镇"): "新蒲新区",
    ("遵义市", "喇叭镇"): "新蒲新区",
    ("遵义市", "新中街道"): "新蒲新区",
    ("遵义市", "新舟镇"): "新蒲新区",
    ("遵义市", "新蒲街道"): "新蒲新区",
    ("遵义市", "礼仪街道"): "新蒲新区",
    ("贵阳市", "党武街道"): "贵安新区",
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_points(source: Path) -> list[dict[str, str]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ValueError("调查点表没有数据行") from exc

    headers: list[str] = []
    for raw in raw_headers:
        header = HEADER_ALIASES.get(_clean(raw))
        if header is None and _clean(raw):
            raise ValueError(f"调查点表存在未知表头：{raw}")
        headers.append(header or "")

    missing = [name for name in REQUIRED_HEADERS if name not in headers]
    if missing:
        raise ValueError(f"调查点表缺少表头：{'、'.join(missing)}")
    if "province" not in headers:
        headers.append("province")

    points: list[dict[str, str]] = []
    seen: set[tuple[str, ...]] = set()
    for line_no, row in enumerate(rows, start=2):
        values = {name: _clean(row[idx]) if idx < len(row) else "" for idx, name in enumerate(headers)}
        if not any(values.values()):
            continue
        if "province" not in values or not values["province"]:
            values["province"] = DEFAULT_PROVINCE

        override_key = (values["city"], values["township"])
        if override_key in COUNTY_OVERRIDES:
            values["county"] = COUNTY_OVERRIDES[override_key]

        empty = [name for name in REGION_FIELDS if not values[name]]
        if empty:
            raise ValueError(f"第 {line_no} 行缺少字段：{'、'.join(empty)}")

        key = tuple(values[name] for name in REGION_FIELDS)
        if key in seen:
            raise ValueError(f"第 {line_no} 行重复：{' / '.join(key)}")
        seen.add(key)
        points.append({name: values[name] for name in REGION_FIELDS})

    if not points:
        raise ValueError("调查点表没有有效数据")
    return points


def convert_region_points(source: Path, output: Path) -> int:
    points = _read_points(source)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(points, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return len(points)


def main() -> None:
    parser = argparse.ArgumentParser(description="转换调查点 Excel 为 region_points.json")
    parser.add_argument("source", type=Path, help="调查点 Excel 路径")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT, help="输出 JSON 路径")
    args = parser.parse_args()
    count = convert_region_points(args.source, args.output)
    provinces = sorted({point["province"] for point in json.loads(args.output.read_text(encoding="utf-8"))})
    print(f"已生成 {args.output}：{count} 个调查点，{len(provinces)} 个省份")


if __name__ == "__main__":
    main()
