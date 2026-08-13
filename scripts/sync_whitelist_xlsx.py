# -*- coding: utf-8 -*-
'''DEPRECATED：docs/权限表.xlsx <-> backend/data/whitelist.db 同步。

自 PRD 权限系统改造（2026-08-13）起，whitelist.db 是实时唯一事实源，
区县/市级业务管理员在网页自行维护名单，xlsx 降级为初始导入模板 + 导出存档物。

本脚本仅保留用于：一次性初始导入 / 灾难恢复。
日常同步禁止再跑（旧 xlsx 会覆盖线上新数据）！如需恢复：先备份 whitelist.db。
'''
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / 'backend'))

from app.persistence import whitelist_db  # noqa: E402

DEFAULT_XLSX = PROJECT_ROOT / 'docs' / '权限表.xlsx'
PROTECTED_PHONES = {'13985000001', '13985000002', '13985000003', '13985000004'}

# Header tokens that identify each sheet type on row 2.
ENUMERATOR_HEADER = ('省', '市', '县', '调查小区', '姓名', '联系电话', '管理员层级')
MANAGER_HEADER = ('省', '市', '县', '姓名', '联系电话', '管理员层级', '备注')


def norm(s):
    if s is None:
        return ''
    s = str(s).strip()
    try:
        f = float(s)
        if f == int(f) and len(str(int(f))) >= 7:
            return str(int(f))
    except (ValueError, OverflowError):
        pass
    return s


def strip_inner(s):
    return ' '.join(s.split())


def detect_sheet_kind(ws):
    '''Return " enumerator" / "manager" / None by reading row 2 headers.'''
    row2 = [ws.cell(row=2, column=c).value for c in range(1, 9)]
    head = tuple(norm(v) for v in row2)
    if head[:len(ENUMERATOR_HEADER)] == ENUMERATOR_HEADER:
        return 'enumerator'
    if head[:len(MANAGER_HEADER)] == MANAGER_HEADER:
        return 'manager'
    return None


def read_enumerator(ws):
    rows = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
        if not any(vals):
            continue
        prov, city, county, community, name, phone, admin_level, remark = [norm(x) for x in vals]
        if not phone:
            continue
        rows.append({
            'phone': phone,
            'name': strip_inner(name),
            'province': prov,
            'city': city,
            'county': county,
            'community': community,
            'admin_level': admin_level or '调查员',
            'remark': remark or '',
        })
    return rows


def read_manager(ws):
    rows = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 8)]
        if not any(vals):
            continue
        prov, city, county, name, phone, admin_level, remark = [norm(x) for x in vals]
        if not phone:
            continue
        rows.append({
            'phone': phone,
            'name': strip_inner(name),
            'province': prov,
            'city': city,
            'county': county,
            'community': '',
            'admin_level': admin_level or '调查员',
            'remark': remark or '',
        })
    return rows


def main():
    print('WARNING: 本脚本已 DEPRECATED（PRD 权限系统改造），仅用于初始导入/恢复；日常名单维护请走网页。')
    ap = argparse.ArgumentParser(description='[DEPRECATED] 同步 权限表.xlsx 与 whitelist.db（仅初始导入/恢复）')
    ap.add_argument('--xlsx', type=Path, default=DEFAULT_XLSX, help='权限表 .xlsx 路径')
    ap.add_argument('--dry-run', action='store_true', help='只打印变更，不写入')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(str(args.xlsx), data_only=True)
    enumerators = []
    managers = []
    for ws in wb.worksheets:
        kind = detect_sheet_kind(ws)
        if kind == 'enumerator':
            enumerators = read_enumerator(ws)
        elif kind == 'manager':
            managers = read_manager(ws)

    xls_phones: set[str] = set()
    inserted = updated = 0
    for rec in enumerators + managers:
        xls_phones.add(rec['phone'])
        if args.dry_run:
            existing = whitelist_db.get_user(rec['phone']) is not None
            action = 'updated' if existing else 'inserted'
        else:
            action = whitelist_db.upsert(rec)
        if action == 'inserted':
            inserted += 1
        else:
            updated += 1

    deleted = 0
    db_rows = whitelist_db.list_all(active_only=True)
    for r in db_rows:
        if r['phone'] not in xls_phones and r['phone'] not in PROTECTED_PHONES:
            if not args.dry_run:
                whitelist_db.delete(r['phone'], soft=True)
            deleted += 1

    print(f'XLSX: {args.xlsx}')
    print(f'  调查员: {len(enumerators)} rows')
    print(f'  管理人员: {len(managers)} rows')
    print(f'  DB upsert: inserted={inserted}, updated={updated}')
    print(f'  DB soft-deleted: {deleted}')
    print(f'  Total phones in XLS: {len(xls_phones)}')
    if args.dry_run:
        print('  (dry-run, no DB write)')


if __name__ == '__main__':
    main()
