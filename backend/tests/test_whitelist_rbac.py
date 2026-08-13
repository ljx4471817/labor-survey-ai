# -*- coding: utf-8 -*-
"""PRD 权限系统改造：纯函数 + 鉴权依赖 + 白名单 CRUD 越权 + batch/enable + 导出 + publish 范围。"""
from __future__ import annotations

from io import BytesIO

import pytest
from fastapi import HTTPException

from app.api import quiz_admin as quiz_admin_api
from app.api import whitelist_admin as wl_api
from app.core.constants import AdminLevel, SysRole
from app.infra import auth as auth_mod
from app.models.schemas import BatchDisableRequest, WhitelistEntry
from app.models.schemas.quiz_admin import PublishRequest
from app.persistence import quiz_db, whitelist_db as wl


@pytest.fixture
def tmp_db(monkeypatch, tmp_path):
    """临时 SQLite + monkeypatch 全局 _conn / DB_PATH（避免污染真实白名单）。"""
    db_file = tmp_path / "whitelist.db"
    monkeypatch.setattr(wl, "DB_PATH", db_file)
    monkeypatch.setattr(wl, "_conn", None)
    yield db_file
    monkeypatch.setattr(wl, "_conn", None)


def _user(role, level, province="贵州省", city="贵阳市", county="南明区"):
    return {
        "phone": "13900000001", "name": "操作员", "sys_role": role,
        "admin_level": level, "province": province, "city": city, "county": county,
        "active": 1,
    }


def _entry(phone, name="测试", admin_level="调查员", province="贵州省", city="贵阳市", county="南明区", sys_role=None):
    return WhitelistEntry(
        phone=phone, name=name, province=province, city=city, county=county,
        admin_level=admin_level, sys_role=sys_role,
    )


def _upsert(phone, name="测试", province="贵州省", city="贵阳市", county="南明区",
            admin_level="调查员", sys_role=None, township="", community=""):
    wl.upsert({
        "phone": phone, "name": name, "province": province, "city": city,
        "county": county, "township": township, "community": community,
        "admin_level": admin_level,
        **({"sys_role": sys_role} if sys_role else {}),
    })


# --- 纯函数 -------------------------------------------------------------------

def test_region_scope_system_admin_unlimited():
    assert auth_mod.region_scope(_user(SysRole.SYSTEM_ADMIN.value, AdminLevel.ENUMERATOR.value)) is None


def test_region_scope_by_level():
    assert auth_mod.region_scope(_user(SysRole.BUSINESS_ADMIN.value, AdminLevel.PROVINCE.value)) == ("贵州省", "", "")
    assert auth_mod.region_scope(_user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)) == ("贵州省", "贵阳市", "")
    assert auth_mod.region_scope(_user(SysRole.BUSINESS_ADMIN.value, AdminLevel.DISTRICT.value)) == ("贵州省", "贵阳市", "南明区")
    # 普通用户/调查员按最小范围（本县）兜底
    assert auth_mod.region_scope(_user(SysRole.USER.value, AdminLevel.ENUMERATOR.value)) == ("贵州省", "贵阳市", "南明区")


def test_in_scope_matrix():
    sa = _user(SysRole.SYSTEM_ADMIN.value, AdminLevel.ENUMERATOR.value)
    prov = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.PROVINCE.value)
    city = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    county = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.DISTRICT.value)
    t_gz = {"province": "贵州省", "city": "贵阳市", "county": "南明区"}
    t_zy = {"province": "贵州省", "city": "遵义市", "county": "汇川区"}
    t_out = {"province": "四川省", "city": "成都市", "county": "武侯区"}
    assert auth_mod.in_scope(sa, t_zy) and auth_mod.in_scope(sa, t_out)
    assert auth_mod.in_scope(prov, t_gz) and auth_mod.in_scope(prov, t_zy)
    assert not auth_mod.in_scope(prov, t_out)
    assert auth_mod.in_scope(city, t_gz) and not auth_mod.in_scope(city, t_zy)
    assert auth_mod.in_scope(county, t_gz) and not auth_mod.in_scope(county, t_zy)
    assert auth_mod.in_scope(city, None)  # 目标缺失兜底放行（调用方负责 404）


def test_allowed_admin_levels_matrix():
    sa = _user(SysRole.SYSTEM_ADMIN.value, AdminLevel.ENUMERATOR.value)
    prov = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.PROVINCE.value)
    city = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    county = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.DISTRICT.value)
    assert set(auth_mod.allowed_admin_levels(sa)) == set(AdminLevel.values())
    assert set(auth_mod.allowed_admin_levels(prov)) == {"市级", "区县", "调查员"}
    assert set(auth_mod.allowed_admin_levels(city)) == {"区县", "调查员"}
    assert set(auth_mod.allowed_admin_levels(county)) == {"调查员"}
    assert auth_mod.allowed_admin_levels(_user(SysRole.USER.value, AdminLevel.ENUMERATOR.value)) == ()


def test_is_protected_phone():
    for p in ("13985000001", "13985000002", "13985000003", "13985000004"):
        assert auth_mod.is_protected_phone(p)
    assert not auth_mod.is_protected_phone("13900000000")


# --- require_user active 校验 ---------------------------------------------------

def test_require_user_checks_active(monkeypatch):
    token, _ = auth_mod.sign_token("13985000001")
    monkeypatch.setattr(auth_mod, "load_whitelist", lambda: frozenset({"13985000001"}))
    assert auth_mod.require_user("Bearer " + token) == "13985000001"
    # 停用后：即使 token 未过期也 401（离职账号即时失效）
    monkeypatch.setattr(auth_mod, "load_whitelist", lambda: frozenset())
    with pytest.raises(HTTPException) as e:
        auth_mod.require_user("Bearer " + token)
    assert e.value.status_code == 401


# --- CRUD 越权 -------------------------------------------------------------------

def test_district_cannot_create_outside_county(tmp_db):
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.DISTRICT.value, county="南明区")
    with pytest.raises(HTTPException) as e:
        wl_api.create_whitelist(_entry("13800000001", county="云岩区"), user=actor)
    assert e.value.status_code == 403


def test_district_cannot_set_district_level(tmp_db):
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.DISTRICT.value, county="南明区")
    with pytest.raises(HTTPException) as e:
        wl_api.create_whitelist(_entry("13800000001", admin_level="区县", county="南明区"), user=actor)
    assert e.value.status_code == 403


def test_business_admin_sys_role_forced(tmp_db):
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    # 业务管理员 body 里塞 sys_role=系统管理员 -> 忽略，按 admin_level 推导
    wl_api.create_whitelist(_entry("13800000001", admin_level="调查员", sys_role="系统管理员"), user=actor)
    assert wl.get_user_any("13800000001")["sys_role"] == "普通用户"
    # 业务管理员建区县管理岗 -> 推导为业务管理员
    wl_api.create_whitelist(_entry("13800000002", admin_level="区县"), user=actor)
    assert wl.get_user_any("13800000002")["sys_role"] == "业务管理员"


def test_business_admin_cannot_touch_protected_phone(tmp_db):
    _upsert("13985000001", admin_level="调查员")
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    with pytest.raises(HTTPException) as e:
        wl_api.remove_whitelist("13985000001", user=actor)
    assert e.value.status_code == 403
    with pytest.raises(HTTPException) as e:
        wl_api.update_whitelist("13985000001", _entry("13985000001", name="改"), user=actor)
    assert e.value.status_code == 403


def test_business_admin_cannot_touch_system_admin_account(tmp_db):
    _upsert("13985000001", admin_level="调查员", sys_role="系统管理员")
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    with pytest.raises(HTTPException) as e:
        wl_api.remove_whitelist("13985000001", user=actor)
    assert e.value.status_code == 403
    with pytest.raises(HTTPException) as e:
        wl_api.enable_whitelist("13985000001", user=actor)
    assert e.value.status_code == 403


def test_district_crud_in_county_ok(tmp_db):
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.DISTRICT.value, county="南明区")
    assert wl_api.create_whitelist(_entry("13800000001", county="南明区"), user=actor)["ok"]
    assert wl_api.update_whitelist("13800000001", _entry("13800000001", name="新名", county="南明区"), user=actor)["ok"]
    assert wl.get_user("13800000001")["name"] == "新名"
    assert wl_api.enable_whitelist("13800000001", user=actor)["ok"]
    assert wl_api.remove_whitelist("13800000001", user=actor)["ok"]
    assert wl.get_user("13800000001") is None


def test_system_admin_global_crud(tmp_db):
    actor = _user(SysRole.SYSTEM_ADMIN.value, AdminLevel.ENUMERATOR.value)
    r = wl_api.create_whitelist(_entry("13800000001", county="遵义市", city="遵义市", sys_role="业务管理员"), user=actor)
    assert r["ok"]
    assert wl.get_user_any("13800000001")["sys_role"] == "业务管理员"


def test_list_scoped(tmp_db):
    _upsert("13800000001", county="南明区")
    _upsert("13800000002", city="遵义市", county="汇川区")
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    items = wl_api.list_whitelist(user=actor)["items"]
    assert [i["phone"] for i in items] == ["13800000001"]
    sa = _user(SysRole.SYSTEM_ADMIN.value, AdminLevel.ENUMERATOR.value)
    assert len(wl_api.list_whitelist(user=sa)["items"]) == 2


# --- batch-disable / enable 语义 --------------------------------------------------

def test_batch_disable_partial_skip(tmp_db):
    _upsert("13800000001", county="南明区")
    _upsert("13800000002", city="遵义市", county="汇川区")
    _upsert("13985000001", county="南明区")
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    res = wl_api.batch_disable(
        BatchDisableRequest(phones=["13800000001", "13800000002", "13985000001", "19999999999"]),
        user=actor,
    )
    assert res["disabled"] == 1
    reasons = {s["phone"]: s["reason"] for s in res["skipped"]}
    assert reasons == {"138****0002": "out_of_scope", "139****0001": "protected", "199****9999": "not_found"}
    assert wl.get_user("13800000001") is None


def test_put_does_not_reactivate_and_enable_recovers(tmp_db):
    _upsert("13800000001")
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    wl.delete("13800000001", soft=True)
    wl_api.update_whitelist("13800000001", _entry("13800000001", name="改名"), user=actor)
    u = wl.get_user_any("13800000001")
    assert u["name"] == "改名" and u["active"] == 0  # PUT 不复活
    wl_api.enable_whitelist("13800000001", user=actor)
    assert wl.get_user("13800000001") is not None


# --- 导出 --------------------------------------------------------------------------

def test_export_xlsx_two_sheets(tmp_db):
    _upsert("13800000001", name="调查员甲", admin_level="调查员")
    _upsert("13800000002", name="区县乙", admin_level="区县")
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)
    resp = wl_api.export_whitelist(user=actor)
    assert "spreadsheetml" in resp.media_type
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(resp.body))
    assert wb.sheetnames == ["调查员", "管理人员"]
    assert [c.value for c in wb["调查员"][1]] == ["省", "市", "县", "调查小区", "姓名", "联系电话", "管理员层级", "备注"]
    assert [c.value for c in wb["管理人员"][1]] == ["省", "市", "县", "姓名", "联系电话", "管理员层级", "备注"]
    assert wb["调查员"].max_row == 2 and wb["管理人员"].max_row == 2


def test_export_csv_for_district(tmp_db):
    _upsert("13800000001", name="调查员甲", admin_level="调查员")
    _upsert("13800000002", city="遵义市", county="汇川区", admin_level="调查员")
    actor = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.DISTRICT.value, county="南明区")
    resp = wl_api.export_whitelist(user=actor)
    assert "text/csv" in resp.media_type
    text = resp.body.decode("utf-8")
    assert "省,市,县,调查小区,姓名,联系电话,管理员层级,备注" in text
    assert "13800000001" in text and "13800000002" not in text  # 区外不导出


# --- quiz publish 范围校验 -----------------------------------------------------------

def test_quiz_publish_out_of_scope_422(monkeypatch):
    admin = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.CITY.value)

    def fake_get_user(p):
        if p == admin["phone"]:
            return admin
        if p == "13800000002":
            return {"phone": p, "province": "贵州省", "city": "遵义市", "county": "汇川区",
                    "admin_level": "调查员", "sys_role": "普通用户", "active": 1}
        return {"phone": p, "province": "贵州省", "city": "贵阳市", "county": "南明区",
                "admin_level": "调查员", "sys_role": "普通用户", "active": 1}

    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.get_user", fake_get_user)
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.get_user_any", fake_get_user)
    monkeypatch.setattr(quiz_db, "get_quiz", lambda q: {"id": q, "status": "draft"})
    with pytest.raises(HTTPException) as e:
        quiz_admin_api.quiz_publish(
            PublishRequest(quiz_id="Q1", targets=["13800000001", "13800000002"], action="publish"),
            phone="13900000001",
        )
    assert e.value.status_code == 422
    assert e.value.detail == ["138****0002"]


def test_quiz_stats_district_scope(monkeypatch):
    """区县业务管理员：完成率只读统计只返回本县行（PRD 矩阵：区县→本县）。"""
    district = _user(SysRole.BUSINESS_ADMIN.value, AdminLevel.DISTRICT.value, county="南明区")
    users = [
        {"phone": "13800000001", "name": "本县甲", "province": "贵州省", "city": "贵阳市", "county": "南明区",
         "admin_level": "调查员", "sys_role": "普通用户", "active": 1},
        {"phone": "13800000002", "name": "外县乙", "province": "贵州省", "city": "贵阳市", "county": "云岩区",
         "admin_level": "调查员", "sys_role": "普通用户", "active": 1},
    ]

    def fake_get_user(p):
        if p == district["phone"]:
            return district
        return next((u for u in users if u["phone"] == p), None)

    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.get_user", fake_get_user)
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.list_all", lambda active_only=False: users)
    monkeypatch.setattr(quiz_db, "get_quiz", lambda q: {"id": q, "month": "2026-08", "title": "测试", "status": "published"})
    monkeypatch.setattr(quiz_db, "list_target_phones", lambda q: ["13800000001", "13800000002"])
    monkeypatch.setattr(quiz_db, "answered_phones", lambda q: set())
    monkeypatch.setattr(quiz_db, "count_answers", lambda q, p: 0)
    monkeypatch.setattr(quiz_db, "count_correct", lambda q, p: 0)
    monkeypatch.setattr(quiz_db, "count_questions", lambda q: 1)
    monkeypatch.setattr(quiz_db, "latest_answer_ts", lambda q, p: None)
    monkeypatch.setattr(quiz_db, "sync_expired", lambda *a, **k: 0)
    monkeypatch.setattr(quiz_db, "cleanup_expired", lambda *a, **k: {"archived": 0})

    st = quiz_admin_api.quiz_stats(
        quiz_id="Q1", region=None, q=None, page=1, page_size=50, phone=district["phone"],
    )
    assert st["total_users"] == 1  # 只统计本县
    assert st["user_details"][0]["county"] == "南明区"
    assert st["by_region"] == [{"region": "南明区", "total": 1, "completed": 0, "rate": 0.0}]


def test_quiz_publish_system_admin_unlimited(monkeypatch):
    sa = _user(SysRole.SYSTEM_ADMIN.value, AdminLevel.ENUMERATOR.value)

    def fake_get_user(p):
        if p == sa["phone"]:
            return sa
        return {"phone": p, "province": "四川省", "city": "成都市", "county": "武侯区",
                "admin_level": "调查员", "sys_role": "普通用户", "active": 1}

    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.get_user", fake_get_user)
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.get_user_any", fake_get_user)
    monkeypatch.setattr(quiz_db, "get_quiz", lambda q: {"id": q, "status": "draft"})
    monkeypatch.setattr(quiz_db, "list_questions", lambda q: [{"selected": 1, "status": "approved"}])
    monkeypatch.setattr(quiz_db, "set_targets", lambda q, phones: len(phones))
    monkeypatch.setattr(quiz_db, "update_quiz", lambda q, **k: None)
    res = quiz_admin_api.quiz_publish(
        PublishRequest(quiz_id="Q1", targets=["13800000001"], action="publish", valid_until=None),
        phone=sa["phone"],
    )
    assert res["ok"] is True
