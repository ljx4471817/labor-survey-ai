# -*- coding: utf-8 -*-
"""quiz API 层测试：可见性 / 判分锁定 / 过期 / 下发 / 统计 / 权限（PRD v3 10.1）。"""
from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO

import pytest
from fastapi import HTTPException, UploadFile

from app.api import quiz as quiz_api
from app.api import quiz_admin as quiz_admin_api
from app.infra import auth as auth_mod
from app.models.schemas.quiz import QuizSubmitRequest
from app.models.schemas.quiz_admin import (
    ExtractRequest,
    KeypointReviewRequest,
    PublishRequest,
    QuestionReviewRequest,
)
from app.persistence import quiz_db

UTC8 = timezone(timedelta(hours=8))


@pytest.fixture
def db(tmp_path, monkeypatch):
    monkeypatch.setattr(quiz_db, "DB_PATH", tmp_path / "quiz_test.db")
    monkeypatch.setattr(quiz_admin_api, "TMP_DIR", tmp_path / "tmp")
    quiz_db.reset_conn()
    yield quiz_db


def _now_iso() -> str:
    return datetime.now(UTC8).isoformat(timespec="seconds")


def _seed_published(db, phone="13800000001", valid_days=7, approved=True):
    qid = db.create_quiz("测试测验", created_by="admin", month="2026-08")
    db.replace_keypoints(qid, [{
        "section": "审核要点",
        "content": "家务劳动者无收入应判为非劳动力",
        "common_error": "误判为就业",
        "source_quote": "将家务劳动者误判为就业是本月常见错误。",
        "suggest_quiz": True,
    }])
    db.replace_questions(qid, [{
        "question": "家务劳动者无收入应判定为？",
        "options": '{"A": "就业人口", "B": "失业人口", "C": "非劳动力", "D": "在职未就业"}',
        "answer": "C",
        "explanation": "根据审核要点应判为非劳动力。",
        "created_by": "admin",
    }])
    if approved:
        for q in db.list_questions(qid):
            db.update_question(q["id"], status="approved", selected=1)
    valid_until = (datetime.now(UTC8) + timedelta(days=valid_days)).isoformat(timespec="seconds")
    db.set_targets(qid, [phone])
    db.update_quiz(qid, status="published", valid_from=_now_iso(), valid_until=valid_until)
    return qid


# --- 调查员端 ---

def test_current_empty_for_non_target(db):
    _seed_published(db)
    data = quiz_api.current(phone="13899999999")
    assert data == {"items": []}


def test_current_for_target_does_not_leak_answer(db):
    qid = _seed_published(db)
    data = quiz_api.current(phone="13800000001")
    assert len(data["items"]) == 1
    q = data["items"][0]["questions"][0]
    assert q["id"].startswith(qid)
    assert "answer" not in q
    assert "explanation" not in q
    assert data["items"][0]["answered"] == 0


def test_submit_flow(db):
    qid = _seed_published(db)
    q = quiz_db.list_questions(qid)[0]
    res = quiz_api.submit(QuizSubmitRequest(quiz_id=qid, q_id=q["id"], selected="C"), phone="13800000001")
    assert res["correct"] is True
    assert res["answer"] == "C"
    assert res["completed"] is True  # 1/1 题全部答完
    assert res["answered"] == 1 and res["total"] == 1
    # 全部答完 → completed True（上一步 total=1 时 completed 应为 True；此处断言兼容）
    # 重新读取 current：已答题带解析
    cur = quiz_api.current(phone="13800000001")
    qv = cur["items"][0]["questions"][0]
    assert qv["answered"] == "C" and qv["correct"] is True
    assert qv["answer"] == "C"


def test_submit_wrong_answer_scores_zero(db):
    qid = _seed_published(db)
    q = quiz_db.list_questions(qid)[0]
    res = quiz_api.submit(QuizSubmitRequest(quiz_id=qid, q_id=q["id"], selected="A"), phone="13800000001")
    assert res["correct"] is False


def test_submit_duplicate_409(db):
    qid = _seed_published(db)
    q = quiz_db.list_questions(qid)[0]
    quiz_api.submit(QuizSubmitRequest(quiz_id=qid, q_id=q["id"], selected="C"), phone="13800000001")
    with pytest.raises(HTTPException) as e:
        quiz_api.submit(QuizSubmitRequest(quiz_id=qid, q_id=q["id"], selected="A"), phone="13800000001")
    assert e.value.status_code == 409


def test_submit_expired_409(db):
    qid = _seed_published(db, valid_days=-1)
    q = quiz_db.list_questions(qid)[0]
    with pytest.raises(HTTPException) as e:
        quiz_api.submit(QuizSubmitRequest(quiz_id=qid, q_id=q["id"], selected="C"), phone="13800000001")
    assert e.value.status_code == 409


def test_submit_non_target_403(db):
    qid = _seed_published(db)
    q = quiz_db.list_questions(qid)[0]
    with pytest.raises(HTTPException) as e:
        quiz_api.submit(QuizSubmitRequest(quiz_id=qid, q_id=q["id"], selected="C"), phone="13899999999")
    assert e.value.status_code == 403


def test_submit_invalid_option_422(db):
    qid = _seed_published(db)
    q = quiz_db.list_questions(qid)[0]
    with pytest.raises(HTTPException) as e:
        quiz_api.submit(QuizSubmitRequest(quiz_id=qid, q_id=q["id"], selected="E"), phone="13800000001")
    assert e.value.status_code == 422


def test_history_lists_participated(db):
    qid = _seed_published(db)
    assert quiz_api.history(month=None, page=1, page_size=10, phone="13800000001")["total"] == 1
    assert quiz_api.history(month=None, page=1, page_size=10, phone="13899999999")["total"] == 0


def test_faq_detail(db):
    d = quiz_api.faq_detail("001", phone="13800000001")
    assert d["id"] == "001"
    assert d["question"]
    with pytest.raises(HTTPException) as e:
        quiz_api.faq_detail("999", phone="13800000001")
    assert e.value.status_code == 404


# --- 管理端 ---

def test_import_allows_multiple_quizzes_same_month(db):
    """多场景改造：同月可重复导入（每次导入 = 独立测验），不再 409。"""
    f1 = UploadFile(filename="a.docx", file=BytesIO(b"PK\x03\x04" + b"\x00" * 100))
    r1 = quiz_admin_api.quiz_import(title="第一套", scene="月度通知", month="2026-08", file=f1, phone="13900000001")
    f2 = UploadFile(filename="b.docx", file=BytesIO(b"PK\x03\x04" + b"\x00" * 100))
    r2 = quiz_admin_api.quiz_import(title="第二套", scene="新员工培训", month="2026-08", file=f2, phone="13900000001")
    assert r1["quiz_id"] != r2["quiz_id"]
    assert quiz_db.get_quiz(r2["quiz_id"])["scene"] == "新员工培训"


def test_import_accepts_docx(db, tmp_path):
    f = UploadFile(filename="通知.docx", file=BytesIO(b"PK\x03\x04" + b"\x00" * 100))
    res = quiz_admin_api.quiz_import(title="9月工作提示测试", scene="月度通知", month="2026-09", file=f, phone="13900000001")
    assert res["import_id"].startswith("IMP")
    assert res["quiz_id"].startswith("Q")
    quiz = quiz_db.get_quiz(res["quiz_id"])
    assert quiz["scene"] == "月度通知" and quiz["month"] == "2026-09" and quiz["title"] == "9月工作提示测试"
    assert quiz_db.get_import(res["import_id"])["status"] == "imported"


def test_keypoint_review(db):
    qid = db.create_quiz("t", created_by="admin", month="2026-08")
    db.replace_keypoints(qid, [{"section": "审核要点", "content": "要点"}])
    kp = db.list_keypoints(qid)[0]
    quiz_admin_api.quiz_keypoint_review(KeypointReviewRequest(keypoint_id=kp["id"], action="approve"), phone="13900000001")
    assert db.get_keypoint(kp["id"])["status"] == "approved"
    quiz_admin_api.quiz_keypoint_review(KeypointReviewRequest(keypoint_id=kp["id"], action="edit", edits={"content": "新内容"}), phone="13900000001")
    assert db.get_keypoint(kp["id"])["content"] == "新内容"


def test_question_review(db):
    qid = db.create_quiz("t", created_by="admin", month="2026-08")
    db.replace_questions(qid, [{"question": "q?", "options": '{"A": "1", "B": "2", "C": "3", "D": "4"}', "answer": "A", "explanation": "e", "created_by": "a"}])
    q = db.list_questions(qid)[0]
    quiz_admin_api.quiz_question_review(QuestionReviewRequest(question_id=q["id"], action="approve"), phone="13900000001")
    assert db.get_question(q["id"])["status"] == "approved"
    quiz_admin_api.quiz_question_review(QuestionReviewRequest(question_id=q["id"], action="reject"), phone="13900000001")
    assert db.get_question(q["id"])["status"] == "draft"


def _patch_quiz_admin_actor(monkeypatch, users=None):
    """把 quiz_admin 的 whitelist 访问指向内存用户（13900000001 = 市级业务管理员）。"""
    admin = {
        "phone": "13900000001", "name": "管理员", "province": "贵州省", "city": "贵阳市",
        "county": "", "admin_level": "市级", "sys_role": "业务管理员", "active": 1,
    }
    targets = users if users is not None else [
        {"phone": "13800000001", "name": "张三", "province": "贵州省", "city": "贵阳市", "county": "南明区", "admin_level": "调查员", "sys_role": "普通用户", "active": 1},
        {"phone": "13800000002", "name": "李四", "province": "贵州省", "city": "贵阳市", "county": "云岩区", "admin_level": "调查员", "sys_role": "普通用户", "active": 1},
        {"phone": "13800000003", "name": "王五", "province": "贵州省", "city": "贵阳市", "county": "观山湖区", "admin_level": "调查员", "sys_role": "普通用户", "active": 1},
    ]
    monkeypatch.setattr(
        "app.api.quiz_admin.whitelist_db.get_user",
        lambda p: admin if p == "13900000001" else next((u for u in targets if u["phone"] == p), None),
    )
    monkeypatch.setattr(
        "app.api.quiz_admin.whitelist_db.get_user_any",
        lambda p: admin if p == "13900000001" else next((u for u in targets if u["phone"] == p), None),
    )
    return admin


def test_publish_requires_approved_questions(db, monkeypatch):
    _patch_quiz_admin_actor(monkeypatch)
    qid = _seed_published(db, approved=False)
    quiz_db.update_quiz(qid, status="draft")  # 重置为草稿
    with pytest.raises(HTTPException) as e:
        quiz_admin_api.quiz_publish(PublishRequest(quiz_id=qid, valid_until=None, targets=["13800000001"], action="publish"), phone="13900000001")
    assert e.value.status_code == 422


def test_publish_success_and_append_remove(db, monkeypatch):
    _patch_quiz_admin_actor(monkeypatch)
    qid = _seed_published(db, approved=True)
    quiz_db.update_quiz(qid, status="draft")
    res = quiz_admin_api.quiz_publish(
        PublishRequest(quiz_id=qid, valid_until=(datetime.now(UTC8) + timedelta(days=7)).isoformat(timespec="seconds"), targets=["13800000001", "13800000002"], action="publish"),
        phone="13900000001",
    )
    assert res["ok"] is True and res["total_users"] == 2
    quiz = quiz_db.get_quiz(qid)
    assert quiz["status"] == "published"
    # append
    r2 = quiz_admin_api.quiz_publish(PublishRequest(quiz_id=qid, targets=["13800000003"], action="append"), phone="13900000001")
    assert r2["added"] == 1
    # remove unanswered（13800000002 未答）
    r3 = quiz_admin_api.quiz_publish(PublishRequest(quiz_id=qid, targets=["13800000002"], action="remove"), phone="13900000001")
    assert r3["removed"] == ["13800000002"]


def test_stats_with_fake_whitelist(db, monkeypatch):
    qid = _seed_published(db)
    users = [
        {"phone": "13800000001", "name": "张三", "province": "贵州省", "city": "贵阳市", "county": "南明区", "admin_level": "调查员", "active": 1},
        {"phone": "13800000002", "name": "李四", "province": "贵州省", "city": "贵阳市", "county": "云岩区", "admin_level": "调查员", "active": 1},
    ]
    monkeypatch.setattr(quiz_db, "list_target_phones", lambda q: ["13800000001", "13800000002"])
    monkeypatch.setattr(quiz_db, "answered_phones", lambda q: {"13800000001"})
    monkeypatch.setattr(quiz_db, "count_answers", lambda q, p: 1 if p == "13800000001" else 0)
    monkeypatch.setattr(quiz_db, "count_correct", lambda q, p: 1 if p == "13800000001" else 0)
    monkeypatch.setattr(quiz_db, "latest_answer_ts", lambda q, p: "2026-08-03T10:00:00+08:00")
    monkeypatch.setattr(quiz_db, "count_questions", lambda q: 1)
    monkeypatch.setattr(quiz_db, "sync_expired", lambda *a, **k: 0)
    monkeypatch.setattr(quiz_db, "cleanup_expired", lambda *a, **k: {"archived": 0})
    admin = {
        "phone": "13900000001", "name": "管理员", "province": "贵州省", "city": "贵阳市",
        "county": "", "admin_level": "市级", "sys_role": "业务管理员", "active": 1,
    }
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.list_all", lambda active_only=False: users)
    monkeypatch.setattr(
        "app.api.quiz_admin.whitelist_db.get_user",
        lambda p: admin if p == "13900000001" else next((u for u in users if u["phone"] == p), None),
    )

    st = quiz_admin_api.quiz_stats(quiz_id=qid, region=None, q=None, page=1, page_size=50, phone="13900000001")
    assert st["total_users"] == 2
    assert st["started"] == 1
    assert st["completed"] == 1
    assert st["completion_rate"] == 0.5
    assert len(st["by_region"]) == 2
    assert len(st["user_details"]) == 2
    assert st["total"] == 2
    by_role = {r["role"]: r for r in st["by_role"]}
    assert by_role["调查员"]["total"] == 2 and by_role["调查员"]["completed"] == 1
    # 搜索姓名过滤明细
    st2 = quiz_admin_api.quiz_stats(quiz_id=qid, region=None, q="张三", page=1, page_size=50, phone="13900000001")
    assert st2["total"] == 1 and st2["user_details"][0]["name"] == "张三"


def test_targets_groups_by_city_and_role(monkeypatch):
    import tempfile
    from pathlib import Path
    _tmp_db = Path(tempfile.gettempdir()) / "test_targets_whitelist.db"
    if _tmp_db.exists():
        _tmp_db.unlink()
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.DB_PATH", _tmp_db)
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db._conn", None)
    users = [
        {"phone": "13800000001", "name": "张三", "city": "贵阳市", "county": "南明区", "admin_level": "调查员", "active": 1},
        {"phone": "13800000002", "name": "李四", "city": "贵阳市", "county": "云岩区", "admin_level": "调查员", "active": 1},
        {"phone": "13900000001", "name": "管理员甲", "city": "贵阳市", "county": "", "admin_level": "市级", "active": 1},
        {"phone": "13900000002", "name": "管理员乙", "city": "遵义市", "county": "汇川区", "admin_level": "区县", "active": 1},
    ]
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.list_all", lambda active_only=True: users)
    res = quiz_admin_api.quiz_targets(q=None, phone="13900000001")
    assert res["total"] == 4
    cities = {c["city"]: c for c in res["cities"]}
    assert set(cities) == {"贵阳市", "遵义市"}
    guiyang_roles = {r["role"]: r for r in cities["贵阳市"]["roles"]}
    assert guiyang_roles["调查员"]["count"] == 2
    assert guiyang_roles["市级"]["count"] == 1
    assert guiyang_roles["调查员"]["label"] == "调查员"
    assert guiyang_roles["市级"]["label"] == "市级管理者"
    # 搜索
    res2 = quiz_admin_api.quiz_targets(q="张三", phone="13900000001")
    assert res2["total"] == 1


def _user(role, level):
    return {
        "admin_level": level, "sys_role": role, "province": "贵州省", "city": "贵阳市",
        "county": "南明区", "phone": "13800000001",
    }


def test_require_quiz_admin_matrix(monkeypatch):
    """require_quiz_admin：系统管理员任意层级放行；业务管理员仅省级/市级。"""
    monkeypatch.setattr(auth_mod, "require_user", lambda authorization: "13800000001")
    cases = [
        ("系统管理员", "调查员", True),
        ("业务管理员", "省级", True),
        ("业务管理员", "市级", True),
        ("业务管理员", "区县", False),
        ("业务管理员", "调查员", False),
        ("普通用户", "市级", False),
    ]
    for role, level, expected_ok in cases:
        monkeypatch.setattr(auth_mod, "get_current_user", lambda phone, _r=role, _l=level: _user(_r, _l))
        if expected_ok:
            assert auth_mod.require_quiz_admin("Bearer x") == "13800000001"
        else:
            with pytest.raises(HTTPException) as e:
                auth_mod.require_quiz_admin("Bearer x")
            assert e.value.status_code == 403


def test_require_quiz_stats_matrix(monkeypatch):
    """require_quiz_stats：业务管理员任意层级（含区县）可看只读统计。"""
    monkeypatch.setattr(auth_mod, "require_user", lambda authorization: "13800000001")
    for role, level, expected_ok in [
        ("系统管理员", "调查员", True),
        ("业务管理员", "区县", True),
        ("业务管理员", "市级", True),
        ("普通用户", "区县", False),
    ]:
        monkeypatch.setattr(auth_mod, "get_current_user", lambda phone, _r=role, _l=level: _user(_r, _l))
        if expected_ok:
            assert auth_mod.require_quiz_stats("Bearer x") == "13800000001"
        else:
            with pytest.raises(HTTPException) as e:
                auth_mod.require_quiz_stats("Bearer x")
            assert e.value.status_code == 403


def test_require_system_admin_matrix(monkeypatch):
    monkeypatch.setattr(auth_mod, "require_user", lambda authorization: "13800000001")
    for role, expected_ok in [("系统管理员", True), ("业务管理员", False), ("普通用户", False)]:
        monkeypatch.setattr(auth_mod, "get_current_user", lambda phone, _r=role: _user(_r, "市级"))
        if expected_ok:
            assert auth_mod.require_system_admin("Bearer x") == "13800000001"
        else:
            with pytest.raises(HTTPException) as e:
                auth_mod.require_system_admin("Bearer x")
            assert e.value.status_code == 403


def test_require_whitelist_admin_returns_user(monkeypatch):
    monkeypatch.setattr(auth_mod, "require_user", lambda authorization: "13800000001")
    monkeypatch.setattr(auth_mod, "get_current_user", lambda phone: _user("业务管理员", "区县"))
    u = auth_mod.require_whitelist_admin("Bearer x")
    assert u["sys_role"] == "业务管理员" and u["admin_level"] == "区县"


def test_import_accepts_supported_exts(db):
    _MAGIC = {
        ".doc": b"\xd0\xcf\x11\xe0" + b"\x00" * 100,
        ".wps": b"\xd0\xcf\x11\xe0" + b"\x00" * 100,
        ".docx": b"PK\x03\x04" + b"\x00" * 100,
        ".pdf": b"%PDF" + b"\x00" * 100,
        ".pptx": b"PK\x03\x04" + b"\x00" * 100,
    }
    for fn in ("通知.doc", "通知.wps", "通知.docx", "培训.pdf", "培训.pptx"):
        ext = "." + fn.rsplit(".", 1)[-1]
        f = UploadFile(filename=fn, file=BytesIO(_MAGIC[ext]))
        res = quiz_admin_api.quiz_import(title="t", scene="月度通知", month="", file=f, phone="13900000001")
        assert res["quiz_id"].startswith("Q"), fn
    # 非法类型拒绝
    f = UploadFile(filename="a.exe", file=BytesIO(b"x"))
    with pytest.raises(HTTPException) as e:
        quiz_admin_api.quiz_import(title="t", scene="月度通知", month="", file=f, phone="13900000001")
    assert e.value.status_code == 400


def test_stats_export_xlsx(db, monkeypatch):
    """导出 Excel：列=8 字段，行数=全部目标（不受分页限制）。"""
    users = [
        {"phone": "13800000001", "name": "张三", "province": "贵州省", "city": "贵阳市", "county": "南明区", "community": "A社区", "admin_level": "调查员", "active": 1},
        {"phone": "13900000001", "name": "管理员甲", "province": "贵州省", "city": "贵阳市", "county": "", "community": "", "admin_level": "市级", "active": 1},
    ]
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.list_all", lambda active_only=False: users)
    monkeypatch.setattr("app.api.quiz_admin.whitelist_db.get_user", lambda p: next(u for u in users if u["phone"] == p))
    monkeypatch.setattr(quiz_db, "list_target_phones", lambda q: ["13800000001", "13900000001"])
    monkeypatch.setattr(quiz_db, "answered_phones", lambda q: {"13800000001"})
    monkeypatch.setattr(quiz_db, "count_answers", lambda q, p: 1 if p == "13800000001" else 0)
    monkeypatch.setattr(quiz_db, "count_correct", lambda q, p: 1 if p == "13800000001" else 0)
    monkeypatch.setattr(quiz_db, "count_questions", lambda q: 1)
    monkeypatch.setattr(quiz_db, "latest_answer_ts", lambda q, p: "2026-08-03T10:00:00+08:00")
    monkeypatch.setattr(quiz_db, "sync_expired", lambda *a, **k: 0)
    monkeypatch.setattr(quiz_db, "cleanup_expired", lambda *a, **k: {"archived": 0})
    qid = quiz_db.create_quiz("测试", created_by="admin", month="2026-08")
    try:
        resp = quiz_admin_api.quiz_stats_export(quiz_id=qid, region=None, q=None, phone="13900000001")
        from openpyxl import load_workbook
        from io import BytesIO

        wb = load_workbook(BytesIO(resp.body))
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert header == ["省", "市", "县", "调查小区", "姓名", "联系电话", "管理员层级", "作答情况"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2
        assert rows[0][4] == "张三"
        assert "已完成" in rows[0][7]
        assert "attachment" in resp.headers["content-disposition"]
    finally:
        for t in ("answers", "targets", "questions", "keypoints", "imports", "quizzes"):
            quiz_db._get_conn().execute(f"DELETE FROM {t}")
        quiz_db._get_conn().commit()
